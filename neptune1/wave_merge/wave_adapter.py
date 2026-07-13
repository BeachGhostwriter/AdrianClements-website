"""
WAVE Adapter
============
DuPont WAVE has no public API. The only integration surface it offers is
"Handling the Reports (Saving and Exporting)" -> Detailed Report exported to
Excel / Word / PDF (source: DuPont WAVE manual).

So this adapter's job is: read the *exported* Detailed Report (we support
Excel/CSV; PDF would need an OCR/table-extraction pass first) and normalize
the fields WAVE calls out into our internal WaveDesign model.

Column names below match a typical WAVE Detailed Report "System Summary"
export. Adjust WAVE_COLUMN_MAP if your export headers differ slightly
between WAVE versions.
"""
import csv
from models import WaveDesign

WAVE_COLUMN_MAP = {
    "project_name": "Project Name",
    "feed_flow_m3d": "Feed Flow (m3/d)",
    "permeate_flow_m3d": "Permeate Flow (m3/d)",
    "concentrate_flow_m3d": "Concentrate Flow (m3/d)",
    "recovery_pct": "System Recovery (%)",
    "num_stages": "Number of Stages",
    "feed_pressure_bar": "Feed Pressure (bar)",
    "feed_conductivity_uScm": "Feed Conductivity (uS/cm)",
    "permeate_conductivity_uScm": "Permeate Conductivity (uS/cm)",
    "concentrate_conductivity_uScm": "Concentrate Conductivity (uS/cm)",
    "specific_energy_kwh_m3": "Specific Energy (kWh/m3)",
    "num_elements": "Total Elements",
}


def load_from_csv(path: str) -> WaveDesign:
    with open(path, newline="", encoding="utf-8") as f:
        row = next(csv.DictReader(f))
    return WaveDesign(
        project_name=row[WAVE_COLUMN_MAP["project_name"]],
        feed_flow_m3d=float(row[WAVE_COLUMN_MAP["feed_flow_m3d"]]),
        permeate_flow_m3d=float(row[WAVE_COLUMN_MAP["permeate_flow_m3d"]]),
        concentrate_flow_m3d=float(row[WAVE_COLUMN_MAP["concentrate_flow_m3d"]]),
        recovery_pct=float(row[WAVE_COLUMN_MAP["recovery_pct"]]),
        num_stages=int(row[WAVE_COLUMN_MAP["num_stages"]]),
        feed_pressure_bar=float(row[WAVE_COLUMN_MAP["feed_pressure_bar"]]),
        feed_conductivity_uScm=float(row[WAVE_COLUMN_MAP["feed_conductivity_uScm"]]),
        permeate_conductivity_uScm=float(row[WAVE_COLUMN_MAP["permeate_conductivity_uScm"]]),
        concentrate_conductivity_uScm=float(row[WAVE_COLUMN_MAP["concentrate_conductivity_uScm"]]),
        specific_energy_kwh_m3=float(row[WAVE_COLUMN_MAP["specific_energy_kwh_m3"]]),
        num_elements=int(row[WAVE_COLUMN_MAP["num_elements"]]),
        source_file=path,
    )


def load_from_master_excel(excel_path) -> WaveDesign:
    """Fallback: build a WaveDesign directly from the client's own RO master
    Excel (used here for the prototype, since we don't have a real WAVE
    export on hand). Mirrors the same normalized schema so downstream code
    is identical regardless of source. `excel_path` can be a filesystem
    path (str) or a file-like object (e.g. a Streamlit UploadedFile)."""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Main INPUT-OUTPUT"]
    source_label = getattr(excel_path, "name", excel_path)
    return WaveDesign(
        project_name="RCDT 3.0 XXL - 3 stage BQ system",
        feed_flow_m3d=ws["F9"].value,
        permeate_flow_m3d=ws["H186"].value,
        concentrate_flow_m3d=ws["H185"].value,
        recovery_pct=ws["F15"].value * 100,
        num_stages=3,
        feed_pressure_bar=ws["F12"].value,
        feed_conductivity_uScm=ws["H18"].value * 1000,
        permeate_conductivity_uScm=ws["J18"].value * 1000,
        concentrate_conductivity_uScm=ws["I18"].value * 1000,
        specific_energy_kwh_m3=None,
        num_elements=None,
        source_file=source_label,
    )
